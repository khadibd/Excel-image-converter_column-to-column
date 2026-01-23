import os
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

INPUT_FILE = r"C:\input\input.xlsx"
OUTPUT_FILE = r"C:\output\output.xlsx"
TEMP_DIR = "temp"


MIN_KB = 200
MAX_KB = 300


def ensure_dirs():
    os.makedirs(TEMP_DIR, exist_ok=True)
    os.makedirs("output", exist_ok=True)


def compress_png_to_jpg(png_path, jpg_path):
    img = PILImage.open(png_path).convert("RGB")

    quality = 95
    while True:
        img.save(jpg_path, "JPEG", quality=quality)
        size_kb = os.path.getsize(jpg_path) / 1024

        if MIN_KB <= size_kb <= MAX_KB or quality <= 30:
            break

        quality -= 5

    return round(size_kb, 1)


def process():
    ensure_dirs()

    wb = load_workbook(INPUT_FILE)
    ws = wb.active

    output_wb = Workbook()
    out_ws = output_wb.active

    for row in range(2, ws.max_row + 1):
        png_path = ws[f"F{row}"].value
        if not png_path or not os.path.exists(png_path):
            continue

        jpg_path = os.path.join(TEMP_DIR, f"row_{row}.jpg")

        size = compress_png_to_jpg(png_path, jpg_path)

        # ✅ IMPORTANT: pass PATH, not file object
        img = XLImage(jpg_path)
        img.width = 120
        img.height = 120

        out_ws.add_image(img, f"G{row}")
        out_ws[f"H{row}"] = f"{size} KB"

    # 🔥 SAVE FIRST — DO NOT DELETE TEMP BEFORE THIS
    output_wb.save(OUTPUT_FILE)

    print("✅ Conversion finished successfully")
